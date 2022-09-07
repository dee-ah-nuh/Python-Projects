# -*- coding: utf-8 -*-
"""
Created on Fri Aug 12 09:37:07 2022

@author: BEEMO
"""


from apiclient import discovery, errors
from httplib2 import Http
from oauth2client import file, tools
from oauth2client import  client as c
#from googleapiclient import errors
from datetime import datetime, timedelta
today=datetime.today().strftime('%Y-%m-%d')
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from googleapiclient.discovery import build
# from google_auth_oauthlib.flow import InstalledAppFlow #May be unneeded
# from google.auth.transport.requests import Request
from oauth2client.service_account import ServiceAccountCredentials
import re
import pandas as pd
#import numpy as np
import csv
import requests
import glob
import os.path
import time
import os
username=os.getlogin()
import gspread
import base64
#import email #may be unneeded
import webbrowser
#import csv
import json
#import closeio_api
import unidecode
from closeio_api import Client as CloseIO_API
#from requests.exceptions import ConnectionError
from email.mime.text import MIMEText
#import xlsxwriter
import traceback
import warnings as warning
import database_connector as db_c
#import Start_and_retain

ticY = time.perf_counter()
#Better name is activate_link_from_email
#This gets the most recent ReEngage Nau Export from our sas.api email and downloads the link in it
#It also grabs the auth for gmail, which is used now
def get_emails(credentials_file_path, clientsecret_file_path):    
    # define credentials and client secret file paths
    SCOPES = ['https://mail.google.com/']
    #credentials_file_path = './credentials4/credentials.json'
    #clientsecret_file_path = './credentials4/client_secret.json'    
    store = file.Storage(credentials_file_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(clientsecret_file_path, SCOPES)
        credentials = tools.run_flow(flow, store)
    gmail = build('gmail', 'v1', credentials=credentials)
 
    # request a list of all the messages
    result = gmail.users().messages().list(maxResults=25,userId='me').execute()
    messages = result.get('messages')  
    # messages is a list of dictionaries where each dictionary contains a message id.
 
    # iterate through all the messages
    for msg in messages:
        # Get the message from its id
        txt = gmail.users().messages().get(userId='me', id=msg['id']).execute()      
        # Use try-except to avoid any Errors
        try:
            # Get value of 'payload' from dictionary 'txt'
            payload = txt['payload']
            headers = payload['headers']
            # Look for Subject
            for d in headers:
                if d['name'] == 'Subject':
                    #We're only looking for the export results so ReEngage NAU is the target.We could genericize this with a field/argument
                    #This also only operates on the most recent one. Maybe I could increment to make it so that I grab older stuff
                    #Or add in another if to only take stuff where something equals somethign like only grabbing emails w subject x that contain word y
                    if d['value']=="Your ReEngage NAU export is ready":  
                        # Get the data and decode it with base 64 decoder.
                        body=base64.urlsafe_b64decode(txt.get("payload").get("body").get("data").encode("ASCII")).decode("utf-8")
                        listX=body.split()
                        for i in listX:
                            #We're looking for the link
                            if i[0:4]=='http':
                                #And then we're opening the link (which here downloads the thing)
                                webbrowser.open(i)
                                return gmail
        except:
            print(traceback.format_exc())
    return gmail

#This function initializes the api services for google drive, sheets, and gspread
#It returns these as 3 credential's which will allow usage of these api's when passes
def get_api_services():
    # define credentials and client secret file paths
    credentials_file_path = './credentials2/credentials.json'
    clientsecret_file_path = './credentials2/client_secret.json'
   
    # define scope
    SCOPE = 'https://www.googleapis.com/auth/drive'

    # define store
    store = file.Storage(credentials_file_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = c.flow_from_clientsecrets(clientsecret_file_path, SCOPE)
        credentials = tools.run_flow(flow, store)

    # define API service
    http = credentials.authorize(Http())
    drive = discovery.build('drive', 'v3', http=http)
    sheets = discovery.build('sheets', 'v4', credentials=credentials)
    #Initialize gspread
    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name('./credentials3/client_secret.json', scope)
    client = gspread.authorize(credentials)
    #Return the allowances
    return drive, sheets, client

#This would allow us to get rid of the predownload from zapier and quickly and efficently query on close for only needed data
#But its still not faster than just predownloading and grabbing the data there lol so it remains dormant. AKa self building query driver
def grab(df):
    start=0
    end=200
    rows={}
    df=df.drop_duplicates(subset='EMPLID',keep='first')
    while start<len(df):
        #This grabs everything with a student id thats in the df youre giving to it
        #So it'd be a completely clean left join
        data={#we can only grab 200 things at a time with this call
        "_limit": 200,
        "query": {
            "queries": [
                {
                    "object_type": "lead",
                    "type": "object_type"
                },
                {
                    "queries": [{
                        "queries":[{
                                        "condition": {
                                            "mode": "full_words",
                                            "type": "term",
                                            "values": []
                                        },
                                        "field": {
                                            "custom_field_id": "lcf_8xBr0KeZQNBtZMOChusviGn94cJSoiLTaQVDHxNx7PC",
                                            "type": "custom_field"
                                        },
                                        "type": "field_condition"
                                    }],"type":"and"
                        }
                    ],
                    "type": "and"
                }
            ],
            "type": "and"
        },
        '_fields':{"lead":['id','custom']},
        "sort": []
        }
        #Here we grab in range of 200 len
        for key in df[start:end]['EMPLID']:
            data['query']['queries'][1]['queries'][0]['queries'][0]['condition']['values'].append(str(key))
        resp=api.post('data/search',data=data)
        for lead in resp['data']:
                #we only care about leadid and custom fields here
            rows[lead['id']]=lead['custom']
        #After we use the range we update it to be right
        start=end
        end+=200
    #we then transform the grabbed data into a dataframe with the stipulations we require
    df2=pd.DataFrame.from_dict(rows,orient='index')
    df2.reset_index(inplace=True)
    df2 = df2.rename(columns = {'index':'id'})
    print(df2.dtypes)
    return df2

###This writes to csv a excel file at a given sheet and then at the end saves the wb as endname
###It can take any number of csv,sheet arguments as long as you include both and theyre all being written to the same sheet
###csv,sheet is the format for args
##def writing_to_workbook(wbname,endname,*args):    
##    #We don't always know if the workbook we're tryna get exists or not, so we try to grab it and if that doesnt work we make it
##    try:
##        wb=load_workbook(endname)
##    except:
##        wb=Workbook(endname)
##        wb.save(endname)
##    #These are the lists of sheets and csv's after they get into dataframes we make these as lists to make accessing easier
##    listDF=[]
##    listWS=[]
##    for i in range(0,len(args)):
##        #The even numbers (aka the 1st arguments(0 focused is weird like that)) are the csv's, so we turn them into dataframes and then append to the list of dataframes
##        if i%2==0:
##            listDF.append(pd.read_csv(args[i],index_col=None))
##        #The odd numbers (aka the 2nd args) are the worksheets that we want to be writing to
##        else:
##            listWS.append(args[i])
##    #We know these are the same length and the index's are mpaired unless the user messed up the arguments, and if they did then it'll error which is good
##    #Otherwise we just go through the full lengths
##    for i in range(0,len(listWS)):
##        #We're assigning the value of listWS[i] to a variable because we have to use append later and its the pandas append, not the list append
##        #It matters and broke it when I tried to directly do listWS[i].append(r)
##        x=listWS[i]
##        print(x)
##        with pd.ExcelWriter(endname, mode = "a", engine = "openpyxl", if_sheet_exists = 'replace') as writer:
##          listDF[i].to_excel(writer, sheet_name = listWS[i], index = False)

#This writes to csv a excel file at a given sheet and then at the end saves the wb as endname
#It can take any number of csv,sheet arguments as long as you include both and theyre all being written to the same sheet
def writing_to_workbook(wbname,endname,*args):    
    #csv,sheet
    #We don't always know if the workbook we're tryna get exists or not, so we try to grab it and if that doesnt work we make it
    try:
        wb=load_workbook(wbname)
    except:
        wb=Workbook(wbname)
    #These are the lists of sheets and csv's after they get into dataframes
    #We make these as lists to make accessing easier
    listDF=[]
    listWS=[]
    listWSName=[]
    for i in range(0,len(args)):
        #The even numbers (aka the 1st arguments(0 focused is weird like that))
        #These are the csv's, so we turn them into dataframes and then append to the list of dataframes
        if i%2==0:
            listDF.append(pd.read_csv(args[i],header=None,index_col=None))
        #The odd numbers (aka the 2nd args)
        #These are the worksheets that we want to be writing to
        else:
            #Same thing as with the workbooks, we dont know if these exist already or not so we try grabbing it and if that doesnt work we make it
            try:
                listWS.append(wb[args[i]])
                listWSName.append(args[i])
            except:
                listWS.append(wb.create_sheet(args[i]))
                listWSName.append(args[i])
    #The briliance of the lists is that for any index the pairs are matched up
    #So for listDF[1], that will be the csv argmunet half of the csv,sheet pair, and so we can just use listWS[1] to get that too
    #We know these are the same length unless the user messed up the arguments, and if they did then it'll error which is good
    #Otherwise we just go through the full lengths
    for i in range(0,len(listWS)):
        #We're assigning the value of listWS[i] to a variable because we have to use append later and its the pandas append, not the list append
        #It matters and broke it when I tried to directly do listWS[i].append(r)
        x=listWS[i]
        print(x)
        #So here x is the tab in the excel sheet we want to append to
        #And r is the rth row in the dataframe listDF[i]
        try:
            for r in dataframe_to_rows(listDF[i],index=False,header=None):
                x.append(r)
        except:
            print(traceback.format_exc())
            with pd.ExcelWriter(wbname, mode = "a", engine = "openpyxl", if_sheet_exists = 'replace') as writer:
              listDF[i].to_excel(writer, sheet_name = listWSName[i], index = False)            
    wb.save(endname)


#This returns the most recently created file at a location that matches regex criteria
#Path is the location up to and including the file name, and is generally going to be where the regex is
#Extension is just the file type, nothing too crazy here
def recent_file(path,extension,location):
    #Assign the name
    files = glob.glob(path + extension)
    #Get the path for the location'th most recently created thing matching
    max_file=sorted(files,key=os.path.getctime)[-location]
    #Return that as a text string
    return max_file

#This writes a dataframe to a google sheet
#Wbname is the sheet name, sheetname is the tab name
#Df is the data frame, gsrpead is the gspread authorization (usually going to be client)
def df_to_sheet(wbname,sheetname,df,gspread,*arg):
    #Open shet
    df=df.fillna('')
    wb=gspread.open(wbname)
    #Open tab
    ws=wb.worksheet(sheetname)
    if arg:
        if arg[0]==1:
            ws.clear()
        if arg[0]==2:
            ws.append_rows(df.values.tolist())
            return
    #Write the values
    ws.update([df.columns.values.tolist()] + df.values.tolist())

#This is so I don't have a billion versions of the same exact process in the doing the dropped and added stuff
#Currently unused
def cleanMerge(merge1,merge2,*args):  
    df=pd.merge(merge1,merge2,on="EMPLID",how='left')
    df=df[df["First_Name_y"].isna()]
    df=df.drop_duplicates(subset=["EMPLID"],keep='first')
    if args:
        df=df.dropna(axis=1,how='all')
    return df

def create_message(sender, to, subject, message_text):
  """Create a message for an email.
  Args:
    sender: Email address of the sender.
    to: Email address of the receiver.
    subject: The subject of the email message.
    message_text: The text of the email message.
  Returns:
    An object containing a base64url encoded email object."""
  message = MIMEText(message_text)
  message['to'] = to
  message['from'] = sender
  message['subject'] = subject
  raw_message = base64.urlsafe_b64encode(message.as_string().encode("utf-8"))
  return {'raw': raw_message.decode("utf-8")}

def send_message(service, user_id, message):
  """Send an email message.
  Args:
    service: Authorized Gmail API service instance.
    user_id: User's email address. The special value "me" can be used to indicate the authenticated user.
    message: Message to be sent.
  Returns:
    Sent Message."""
  message = (service.users().messages().send(userId=user_id, body=message).execute())
  return message

###THis downloads a google sheet tab to csv given google api (here sheets), the sheet id, and the sheet name you want
def download_sheet_to_csv(sheets_instance, spreadsheet_id,sheet_name):
    #We grab that sheet
    result = sheets_instance.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=sheet_name).execute()
    today=datetime.today().strftime('%Y-%m-%d')
    #This sets the download loacation
    output_file = f'C:/Users/{username}/Downloads/todo/{sheet_name} {today}.csv'
    #We iterate across the result1 and write all its values into the csv
    with open(output_file, 'w',newline='') as f:
        writer = csv.writer(f)
        writer.writerows(result.get('values'))
    f.close()
    #We let us know that its done
    print(f'Successfully downloaded {sheet_name} {today}.csv')

#THis gets the id for any google sheet given a drive api and its name
def get_spreadsheet_id(api_service,spreadsheet_name):
    results = []
    page_token = None
    while True:
        try:
            param = {'q': 'mimeType="application/vnd.google-apps.spreadsheet"'}
            if page_token:
                param['pageToken'] = page_token
            files = api_service.files().list(**param).execute()
            results.extend(files.get('files'))
            page_token = files.get('nextPageToken')
            if not page_token:
                break
        except errors.HttpError as error:
            print(traceback.format_exc())
            print(f'An error has occurred: {error}')
            break
    spreadsheet_id = [result.get('id') for result in results if result.get('name') == spreadsheet_name][0]
    return spreadsheet_id

def gsheet_to_df(book,sheet,gspread,*arg):
    wb=gspread.open(book)
    ws=wb.worksheet(sheet)
    data=ws.get_all_values()
    if not arg:
        headers=data.pop(0)
        df=pd.DataFrame(data,columns=headers)
        return df
    else:
        df=pd.DataFrame(data)
        return df



#TODO - by loading config
config = json.load(open('config.json'))

#Initialize stuff
drive, sheets, client = get_api_services()
gmail=get_emails(config['credentials_file_path'], config['clientsecret_file_path'])
pool=db_c.database_connection()
#start_amd retain
#Start_and_retain.main()
#Get the most recent files
cont=False
loop=1
while cont==False:
    try:
        not_enrolled=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_NotEnrolled-{today}*','.csv',1)
        enrolled=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_Enrolled-{today}*','.csv',1)
        billing_dfX=pd.read_excel(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Billing Reports/CloseIO ReEngage Sync.xlsx',sheet_name="Student Starts",usecols='A:J')
        fina=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Financial Aid Reports Automated/current_fina_checklist_items-{today}*','.csv',1)
        fafsa=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Financial Aid Reports Automated/fafsa_loan_docs-{today}*','.csv',1)
        to_do=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/To Do List Report Automated/to_do_list-{today}*','.csv',1)
        holds=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/To Do List Report Automated/holds-{today}*','.csv',1)
        cont=True
    except:
        if loop<12:
            print('Files for today not found, trying again in 10 minutes')
            time.sleep(600)
            loop+=1
        else:            
            not_enrolled=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_NotEnrolled-*','.csv',1)
            enrolled=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_Enrolled-*','.csv',1)
            billing_dfX=pd.read_excel(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Billing Reports/CloseIO ReEngage Sync.xlsx',sheet_name="Student Starts",usecols='A:J')
            fina=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Financial Aid Reports Automated/current_fina_checklist_items-*','.csv',1)
            fafsa=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Financial Aid Reports Automated/fafsa_loan_docs-*','.csv',1)
            to_do=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/To Do List Report Automated/to_do_list-*','.csv',1)
            holds=recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/To Do List Report Automated/holds-*','.csv',1)
            warning.warn('THE FILES ARE NOT UP TO DATE')
            cont=True
#Turn them into data frames
not_enrolled_df=pd.read_csv(not_enrolled)
not_enrolled_df.to_sql('not_enrolled_daily',con=pool,index=False,if_exists='replace')
enrolled_df=pd.read_csv(enrolled)
enrolled_df.to_sql('enrolled_daily',con=pool,index=False,if_exists='replace')
billing_dfX=billing_dfX[billing_dfX['EMG Excluded'] =='OPRE Billed']
billing_df=billing_dfX[["ID"]]
print("Dataframes written")

#This here just makes the list of programs that we care about for later into a list
my_file = open("listX.txt", "r")
data = my_file.read()
data_into_list = data.split("\n")




#Finaid automation
#Copy all data from fafsa_loan_docs to https://docs.google.com/spreadsheets/d/1wWO7f1Ck2B89HUx3uh_ytIKhYpdNZIccIJ7o_QfAOAc/edit#gid=887061316
#FPaste this data to the fafsa datasheet in the nau pipeline (What we just went to)
#FAFSA Datasheet
#ReEngage NAU Pipeline
fafsa_df=pd.read_csv(fafsa,on_bad_lines='skip')
fafsa_df.to_sql('fafsa_daily',con=pool,index=False,if_exists='replace')
df_to_sheet("ReEngage NAU Pipeline","FAFSA Datasheet",fafsa_df,client,1)
df_to_sheet("Clearing Checklist Rest","FAFSA",fafsa_df,client,1)
#Go into holds- and delete all where column L=="N"
#Sort by last name and then sort to EMPLID (?)
#Delete columns A:N in holds detail in nau pipeline
#And then paste the everything from what you just made in holds
#Holds Detail
holds_df=pd.read_csv(holds,on_bad_lines='skip')
holds_df.to_sql('holds_daily',con=pool,index=False,if_exists='replace')
holds_df=holds_df[holds_df["ACTIVE_STUDENT_INDICATOR"]!="N"]
holds_df=holds_df.sort_values(by=['EMPLID'])
holds_df=holds_df.sort_values(by=['LAST_NAME'])
df_to_sheet("ReEngage NAU Pipeline","Holds Detail",holds_df,client,1)
df_to_sheet("Clearing Checklist Rest","Holds",holds_df,client,1)
#In to_do_list- delete all where col O (ACTIVE_STUDENT_INDICATOR) == N
to_do_df=pd.read_csv(to_do,on_bad_lines='skip')
to_do_df.to_sql('to_do_daily',con=pool,index=False,if_exists='replace')
to_do_df=to_do_df[to_do_df["ACTIVE_STUDENT_INDICATOR"]!="N"]
#To do drop the last 3 columns: ADMIN_FUNCTION  MAX_FINANCIAL_AID_YEAR  ACTIVE_STUDENT_INDICATOR
to_do_df=to_do_df.drop(columns=['ADMIN_FUNCTION', 'MAX_FINANCIAL_AID_YEAR', 'ACTIVE_STUDENT_INDICATOR'])
##Go into current_fina_checklist_items and delete all whhere
#col I (CHKLI_ITEM_STATUS_LDESC) ==Completed or = Waived
fina_df=pd.read_csv(fina,on_bad_lines='skip')
fina_df=fina_df[(fina_df["CHKLI_ITEM_STATUS_LDESC"]!="Completed") & (fina_df["CHKLI_ITEM_STATUS_LDESC"]!="Waived")]
fina_df.columns=list(to_do_df)
#Append this to holds_df
df_merge=pd.concat([to_do_df,fina_df],ignore_index=True)
#sort the merged one by emplid then last anem
df_merge=df_merge.sort_values(by=['EMPLID'])
df_merge=df_merge.sort_values(by=['LAST_NAME'])
#Paste this in To DO List Detail
df_to_sheet("ReEngage NAU Pipeline","ToDoList Detail",df_merge,client,1)
df_to_sheet("Clearing Checklist Rest","To Do List",df_merge,client,1)
print("Finaid automation complete")


#Clearing checklist primary
ssid=get_spreadsheet_id(drive,"Clearing Checklist Rest")
download_sheet_to_csv(sheets, ssid,"Important Info")
zzz=client.open("Clearing Checklist Rest")
opre=zzz.worksheet("OPRE Use")
sub=zzz.worksheet("Sub Use")
opreVal=opre.get("A2:E")
subVal=sub.get("A2:E")

#def send_emails_to_Opre():
    #This is to notify if theres an opre
for i in opreVal:
    #print(i[4])
    #i4 is the datedif
    if int(i[4])>=1 and int(i[4])<40000:
        c=create_message("me","jtallen@myfootpath.com",'There is a continuing OPRE','There is a continuing OPRE')
        send_message(gmail,"me",c)
        print("Yeah")
#This is to notify the gs to follow up after 3 days
for i in subVal:
    #I4 is the datedif, while I3 is the gs email
    if int(i[4])==3:
        if i[3]=='crystal.merritte@nau.edu':
            rn='cmerritte@myfootpath.com'
        elif i[3]=='ravenia.gant@nau.edu':
            rn='rgant@myfootpath.com'
        elif i[3]=='Vernon.Eby@nau.edu':
            rn='veby@myfootpath.com'
        elif i[3]=='vernon.rucker@nau.edu':
            rn='vrucker@myfootpath.com'
        elif i[3]=='Erik.Van-Conant@nau.edu':
            rn='evanconant@myfootpath.com'
        else:
            print(i[3])
            rn=i[3]
        c=create_message("me",rn,'3 Day Alert',f'This is a reminder to follow up with {i[1]}, as it has been 3 days since submission')
        send_message(gmail,"me",c)
        #print("Yup")
print("Clearing checklist email check complete")

#send_emails_to_Opre()

#doing the dropped and added stuff
#what an ugly mess, but at least its nicer looking now
#This is about finding students who either have less classes than yesterday or have more classes than yesterday
first_day_df=pd.read_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_Enrolled-2022-01-07 06-07-15 AM.csv')
first_not_df=pd.read_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_NotEnrolled-2022-01-07 06-07-19 AM.csv')
yesterday_df=pd.read_csv(recent_file(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/ReEngage NAU Reports/Registrar Reports Automated/Group_Enrolled*','.csv',2))
dfDroppedX=yesterday_df[["EMPLID","First_Name"]].groupby(["EMPLID"]).count()
dfDroppedX=dfDroppedX.rename(columns={"First_Name":"Class Count Yesterday"})
dfDropped=pd.merge(yesterday_df,dfDroppedX, on="EMPLID", how='left')
dfAddedX=enrolled_df[["EMPLID","First_Name"]].groupby(["EMPLID"]).count()
dfAddedX=dfAddedX.rename(columns={"First_Name":"Class Count Today"})
dfAdded=pd.merge(enrolled_df,dfAddedX, on="EMPLID", how='left')
dfAdded2=pd.merge(dfAdded,dfDroppedX,on='EMPLID',how='left')
dfDropped2=pd.merge(dfDropped,dfAddedX,on='EMPLID',how='left')
dfAdded2=dfAdded2.drop_duplicates(subset=['EMPLID'],keep='first')
dfDropped2=dfDropped2.drop_duplicates(subset=['EMPLID'],keep='first')
dfAdded2=dfAdded2[(dfAdded2['Class Count Today']>dfAdded2['Class Count Yesterday'])|(dfAdded2['Class Count Yesterday'].isna())]
dfDropped2=dfDropped2[(dfDropped2['Class Count Yesterday']>dfDropped2['Class Count Today'])|(dfDropped2['Class Count Today'].isna())]
dfDropped2.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Operations/Analysis/Find_Drop_Students/{today} Find_Dropped_Classes.csv',index=False)
dfAdded2.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Operations/Analysis/Find_Add_Students/{today} Find_Added_Classes.csv',index=False)
enrolled_df.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Operations/Analysis/Find_Drop_Students/Enrolled_Today.csv',index=False)
yesterday_df.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Operations/Analysis/Find_Drop_Students/Enrolled_Yesterday.csv',index=False)
print("Added/dropped automation complete")


#Paste enrolled to the google sheet ReEngage NAU Pipeline
# https://docs.google.com/spreadsheets/d/1wWO7f1Ck2B89HUx3uh_ytIKhYpdNZIccIJ7o_QfAOAc/edit#gid=887061316
#In the sheet DailyRegDatasheet
df_to_sheet('ReEngage NAU Pipeline',"DailyRegDatasheet",enrolled_df,client)


#Proof of concept 1
#Sandbox stuff, basically unused, could be killed and it wouldnt matter, might break smth unnescesarily tho and i dont wanna deal w that
not_enrolled_df_X=pd.merge(not_enrolled_df,billing_df,left_on="EMPLID",right_on="ID",how='left').drop_duplicates(subset=["EMPLID"],keep='first')
enrolled_df_X=pd.merge(enrolled_df,billing_df,left_on="EMPLID",right_on="ID",how='left').drop_duplicates(subset=["EMPLID"],keep='first')
enrolled_df_X=enrolled_df_X[['EMPLID','First_Name','Last_Name','Active_Student','TERM', 'CAREER','ID']]
not_enrolled_df_X=not_enrolled_df_X[['EMPLID','First_Name','Last_Name','Active_Student','TERM','CAREER','ID']]
billing_dfX['Billed'] = billing_dfX['Billed'].astype(str)
df_to_sheet("Sandbox","not enrolled",not_enrolled_df_X,client,1)
df_to_sheet("Sandbox","enrolled",enrolled_df_X,client,1)
df_to_sheet("Sandbox","billing",billing_dfX,client,1)


#Get close and convert it
#We wait 2 seconds in case the file from get_emails() isnt fully finished downloading
time.sleep(2)
close=recent_file(f'C:/Users/{username}/Downloads/ReEngage NAU leads*','.csv',1)
close_df=pd.read_csv(close)
close_df.to_sql('close_daily',con=pool,index=False,if_exists='replace')
#editing for auto close upload compliance
close_df=close_df.rename(columns={'id':'custom.id','custom.0.2 Assignee.id':'custom.0.2 Assignee'})
close_df=close_df.drop(['custom.5.3.1 1221 Courses','custom.5.3.2 1224 Courses','custom.5.3.3 1227 Courses','custom.5.3.4 1228 Courses','custom.5.3.5 1231 Courses',"custom.5.2.1 Primary Academic Program"],axis=1)

#Rough Draft automation
#Afaik this is not used at all and turning it off might break something so it just on for a 2 second performance hit
listX=['4.1 Application-ReAdmit','4.2 Application-ReAdmit-RegOfc','4.3 Application-Re-Apply','4.4 Application-Graduate']
close_dfRD=close_df[close_df['status_label'].isin(listX)]
close_dfRD['last_lead_status_change_date'].fillna(close_dfRD['date_created'],inplace=True)
close_dfRD['last_lead_status_change_date']=close_dfRD['last_lead_status_change_date'].str.slice(stop=10)
close_dfRD=close_dfRD[['display_name','status_label','custom.0.2 Assignee.name','custom.3.1 Start Date','custom.3.2 Program to Primary Plan','custom.3.3 Campus to Primary Plan','custom.3.4 Residency','custom.5.1 Student ID','last_lead_status_change_date']]
close_dfRD=close_dfRD.rename(columns={'display_name':'Name','custom.0.2 Assignee.name':'GS','custom.3.1 Start Date':'Start Date','custom.3.2 Program to Primary Plan':'Program to Primary Plan','custom.3.4 Residency':'Residency','custom.5.1 Student ID':'Student ID','last_lead_status_change_date':'Application Date'})
df_to_sheet('Rough Draft','Master List',close_dfRD,client,1)

#Left join enrolled & not enrolled with close
join_enrolled=pd.merge(enrolled_df,close_df, left_on='EMPLID', right_on='custom.5.1 Student ID', how='left')
join_not_enrolled=pd.merge(not_enrolled_df,close_df, left_on='EMPLID', right_on='custom.5.1 Student ID', how='left')


#This is really stupid LMAO saves 7 lines and kills code reuse tho :eyes:
#7 LINES JERRY
def cleanTerm(num,name):
    df=dfterm[dfterm["TERM"]==num].groupby(["EMPLID"]).count()
    df=df.rename(columns={"TERM":"custom."+name})
    hat=pd.merge(join_enrolled,df,on="EMPLID",how='left')
    return hat
dfterm=join_enrolled[["EMPLID","TERM"]]
listNum=(1221,1224,1227,1228,1231)
listName=("5.3.1 1221 Courses","5.3.2 1224 Courses","5.3.3 1227 Courses","5.3.4 1228 Courses","5.3.5 1231 Courses")
for i,j in zip(listNum,listName):
    join_enrolled=cleanTerm(i,j)

billed_students_df=pd.read_excel(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Billing Reports/2022-01-31 Billed Students.xlsx')

#We only care about the stuff that wasn't joined on
join_not_enrolled=join_not_enrolled.drop_duplicates(subset=["EMPLID"],keep='first')
#Sandbox stuff, pointless aorn
f2_df=join_not_enrolled[join_not_enrolled["status_label"]=='5.0 Clearing']
s2_df=join_not_enrolled[join_not_enrolled["status_label"]=='6.0 Start & Retain']
m2_df=join_not_enrolled[(join_not_enrolled["status_label"]=='4.1 Application-ReAdmit')|(join_not_enrolled["status_label"]=='4.2 Application-ReAdmit-RegOfc')|(join_not_enrolled["status_label"]=='4.3 Application-Re-Apply')|(join_not_enrolled["status_label"]=='4.4 Application-Graduate')]

not_enrolled_billing=pd.merge(join_not_enrolled,billed_students_df,left_on='EMPLID',right_on='ID',how='left',indicator='join')
not_enrolled_billing=not_enrolled_billing[['EMPLID','First_Name','Last_Name','custom.3.1 Start Date','join','status_label','custom.PS Clearing']]
not_enrolled_new=not_enrolled_billing[not_enrolled_billing['join']=='left_only']
not_enrolled_cont=not_enrolled_billing[not_enrolled_billing['join']=='both']
not_enrolled_new.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Registrar Course Reports/Not Enrolled New/{today} Not Enrolled New.csv')
not_enrolled_cont.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Registrar Course Reports/Not Enrolled Continuing/{today} Not Enrolled Continuing.csv')

#WITH NO MATCHES: Make list of id's where id==na to get the people not currently in close
lost=join_enrolled[join_enrolled['custom.id'].isna()]
lost=lost.drop_duplicates(subset=['EMPLID'])
lost=lost.dropna(axis=1, how='all')
#This is making the Existing Course Count Upload thing, we drop the extras and the na's here
remainX=join_enrolled[join_enrolled['custom.id'].notna()]
remainX=remainX.drop_duplicates(subset=["EMPLID"])
remain=remainX

enrolled_billing=pd.merge(remain,billed_students_df,left_on='EMPLID',right_on='ID',how='left',indicator='join')
enrolled_billing=enrolled_billing[['EMPLID','First_Name','Last_Name','custom.3.1 Start Date','join','status_label','custom.PS Clearing']]
enrolled_new=enrolled_billing[enrolled_billing['join']=='left_only']
enrolled_cont=not_enrolled_billing[enrolled_billing['join']=='both']
enrolled_new.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Registrar Course Reports/Enrolled New/{today} Enrolled New.csv')
enrolled_cont.to_csv(f'C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Registrar Course Reports/Enrolled Continuing/{today} Enrolled Continuing.csv')
                         

#This sends happy birthday to whoever's bday it is
api = CloseIO_API('api_26qZCo6enkWewoyUHQvbGd.4PZwmPyklAo0vfrk4yWGrd')
#We use this format since thats how its most often put into close, but we dont know for sure since its a text field and not a date field
bday=datetime.today().strftime('%m/%d')
#Filtering out everything that starts with todays formatted date
bday_df=remain[remain['custom.1.8 Date of birth'].str.startswith(bday, na=False)]
#Then we go through
for i,j in bday_df.iterrows():
    em=j['primary_contact_primary_email']
    by=j['custom.id']
    #We need the contact id to do the template thing, but thats not in our exported data
    #We just have to fetch it via api call
    z=api.get('contact',params={"lead_id":by})
    cont=z['data'][0]['id']
    try:
        #Send the templated email as vernon#
        r=api.post('/activity/email',data={'contact_id':cont,'lead_id':by,'status':'outbox','template_id':'tmpl_cIuulY5JmvtBJ8wJr4mGXV7fZXCk15kFQzSn44uJziy','to':[em],'sender':"vernon.rucker@nau.edu"})
        print('Birthday Email Sent')
    except Exception as e:
        print(traceback.format_exc())
#George wanted the column names to be pretty
bday_df=bday_df.rename(columns={"custom.5.1 Student ID":"Student ID",'display_name':'Name','custom.3.3 Campus to Primary Plan':'Campus','custom.0.2 Assignee.name':'GS','custom.3.1 Start Date':'Start Date'})
bday_df['Date Sent']=today
bday_df=bday_df[['Name','Student ID','Campus','GS','Start Date','Date Sent']]
#Posting the bday report to the bday report google sheet, the 2 arg makes it so that this appends instead of overwrites
df_to_sheet('Birthday Report','Sheet1',bday_df,client,2)

   
#WE INTERUPT THIS REGULARLY SCHEDULED DEFINING TO BRING YOU THIS BREAKING NEWS
#None of this is really used anymore, but turning it off could break something and idrc enough to problem solve that when i can just let it run for like a 3 second performance time hit on a 5 minute script
f_df=remain[remain["status_label"]=='5.0 Clearing']
s_df=remain[remain["status_label"]=='6.0 Start & Retain']
m_df=remain[(remain["status_label"]=='4.1 Application-ReAdmit')|(remain["status_label"]=='4.2 Application-ReAdmit-RegOfc')|(remain["status_label"]=='4.3 Application-Re-Apply')|(remain["status_label"]=='4.4 Application-Graduate')]
f_df=f_df.append(f2_df)
s_df=s_df.append(s2_df)
m_df=m_df.append(m2_df)
#Sandbox stuff, supposed to limit code reuse. Kinda pointless aorn since sandbox is dead
def doing(f_df,s_df,m_df,client,*args):
    j=0
    listX=[]
    for i in [f_df,s_df,m_df]:
        i=i.rename(columns={"Primary_Acad_Prog":"Primary Academic Program","EMPLID":"5.1 Student ID","custom.5.3.1 1221 Courses":"1221 Term","custom.5.3.2 1224 Courses":"1224 Term","custom.5.3.3 1227 Courses":"1227 Term","custom.5.3.4 1228 Courses":"1228 Term","custom.5.3.5 1231 Courses":"1231 Term"})
        i=i[["5.1 Student ID","First_Name","Last_Name","1221 Term","1224 Term","1227 Term","1228 Term","1231 Term","Primary Academic Program",'custom.0.2 Assignee.name','status_label']]
        df_to_sheet("Sandbox",args[j],i,client,1)
        j+=1
        i['First_Name']=i['First_Name'].str.title()
        i['Last_Name']=i['Last_Name'].str.title()
        i=i[["5.1 Student ID","1221 Term","1224 Term","1227 Term","1228 Term","1231 Term","Primary Academic Program"]]
        listX.append(i)
    return listX
listX=doing(f_df,s_df,m_df,client,"5.0 clearing","6.0 existing","4.x")
close_dfx=close_df.rename(columns={"custom.5.1 Student ID":"5.1 Student ID",'primary_contact_last_name':'Last_Name','primary_contact_first_name':'First_Name'})
close_dfx['First_Name']=close_dfx['First_Name'].str.title()
close_dfx['Last_Name']=close_dfx['Last_Name'].str.title()
f_df=listX[0]
s_df=listX[1]
m_df=listX[2]
c_f=close_dfx[close_dfx["status_label"]=='5.0 Clearing']
c_f=c_f[["custom.id","5.1 Student ID",'status_label',"custom.0.2 Assignee.name","First_Name","Last_Name","custom.0.4.1 Lead Type"]]
c_f=pd.merge(c_f,f_df, on='5.1 Student ID', how='left')
s_f=close_dfx[close_dfx["status_label"]=='6.0 Start & Retain']
s_f=s_f[["custom.id","5.1 Student ID",'status_label',"custom.0.2 Assignee.name","First_Name","Last_Name","custom.0.4.1 Lead Type"]]
s_f=pd.merge(s_f,s_df, on='5.1 Student ID', how='left')
m_f=close_dfx[(close_dfx["status_label"]=='4.1 Application-ReAdmit')|(close_dfx["status_label"]=='4.2 Application-ReAdmit-RegOfc')|(close_dfx["status_label"]=='4.3 Application-Re-Apply')|(close_dfx["status_label"]=='4.4 Application-Graduate')]
m_f=m_f[["custom.id","5.1 Student ID",'status_label',"custom.0.2 Assignee.name","First_Name","Last_Name","custom.0.4.1 Lead Type"]]
m_f=pd.merge(m_f,m_df, on='5.1 Student ID', how='left')

close_dfY=pd.merge(close_df,enrolled_df,left_on='custom.5.1 Student ID',right_on='EMPLID',how='left',indicator='enr')
close_dfY=pd.merge(close_dfY,not_enrolled_df,left_on='custom.5.1 Student ID',right_on='EMPLID',how='left',indicator='ner')
close_dfY=close_dfY[(close_dfY["status_label"]=='4.1 Application-ReAdmit')|(close_dfY["status_label"]=='5.0 Clearing')|(close_dfY["status_label"]=='6.0 Start & Retain')|(close_dfY["status_label"]=='4.2 Application-ReAdmit-RegOfc')|(close_dfY["status_label"]=='4.3 Application-Re-Apply')|(close_dfY["status_label"]=='4.4 Application-Graduate')]
close_dfY=close_dfY[(close_dfY['enr']=='left_only')&(close_dfY['ner']=='left_only')]
close_dfY=close_dfY.dropna(axis=1,how='all')
close_dfY['enr'] = close_dfY.enr.astype(str)
close_dfY['ner'] = close_dfY.ner.astype(str)

df_to_sheet("Sandbox","4.x Close",m_f,client,1)
df_to_sheet("Sandbox","5.0 Close",c_f,client,1)
df_to_sheet("Sandbox","6.0 Close",s_f,client,1)
df_to_sheet("Sandbox","In Close not in EnrolledNot",close_dfY,client,1)


#Getting the count for the relevant quarters. Will need to be updated as the quarters change
#len12xx is just the total count of enrolled students for that quarter, while len12xxO is the count of all the students we get paid for
#The difference there is that we don't get paid for students whose campus is personalized learning, and we dont get paid for specific online programs
#There is a list of these specific online programs stored at listX.txt in google_api_stuff
len1224=remain[remain["custom.5.3.2 1224 Courses"]!=0]["custom.5.3.2 1224 Courses"].count()
len1227=remain[remain["custom.5.3.3 1227 Courses"]!=0]["custom.5.3.3 1227 Courses"].count()
len1224O=remain[(remain["custom.5.3.2 1224 Courses"]>0)&(remain['Student_Campus']!='Personalized Learning')&~((remain["Primary_Acad_Plan"].isin(data_into_list)==True)&(remain['Student_Campus']=='Online'))]
len1227O=remain[(remain["custom.5.3.3 1227 Courses"]>0)&(remain['Student_Campus']!='Personalized Learning')&~((remain["Primary_Acad_Plan"].isin(data_into_list)==True)&(remain['Student_Campus']=='Online'))]
#Renaming to what it is in the original excel sheet since Vernon was insistent
remainX=remain.rename(columns={"Primary_Acad_Prog":"5.2.1 Primary Academic Program","EMPLID":"5.1 Student ID","custom.5.3.1 1221 Courses":"5.3.1 1221 Courses","custom.5.3.2 1224 Courses":"5.3.2 1224 Courses","custom.5.3.3 1227 Courses":"5.3.3 1227 Courses","custom.5.3.4 1228 Courses":"5.3.4 1228 Courses","custom.5.3.5 1231 Courses":"5.3.5 1231 Courses"})
#Drop every column except what we need
remainX=remainX[["5.1 Student ID","5.3.1 1221 Courses","5.3.2 1224 Courses","5.3.3 1227 Courses","5.3.4 1228 Courses","5.3.5 1231 Courses","5.2.1 Primary Academic Program"]]
#Fill the na's with 0 so that the things arent empty since Vernon was opposed to empty datapoints int eh 5.3.x 12xx courses
remainX=remainX.fillna(0)
#This is preparing the data for posting to close
#We drop assignee name because close wouldnt accept an api post with that field, it needs the GS ID instead
#Custom.5.1 Student ID is being dropped because we use EMPLID instead, although they should just be identical
join_enrolled_v=join_enrolled.drop(['custom.0.2 Assignee.name','custom.5.1 Student ID'],axis=1)
#We update primary academic program to match whatever it is in the enrollment sheet
join_enrolled_v=join_enrolled_v.rename(columns={"Primary_Acad_Prog":"custom.5.2.1 Primary Academic Program","EMPLID":"custom.5.1 Student ID"})
#We drop all non custom fields, since we dont want to update anything but those
join_enrolled_v=join_enrolled_v.loc[:,join_enrolled_v.columns.str.startswith('custom.')]
#Dropping the custom from the names since having it in breaks the api post
join_enrolled_v.columns=join_enrolled_v.columns.str.replace('custom.', '',regex=True)
data_into_list = ['PS Continuing','id','PS Transcript', 'PS Clearing', '0.1 Timezone', '0.2 Assignee', '0.3 Priority', '0.4.1 Lead Type', '0.4.2 Lead Type Description', '0.5.1 Prior Major', '0.5.2 Prior Campus', '0.5.3 Prior Academic Career', '1.1 Interested in Online?', '1.2 Have Computer + Internet in Home', '1.3 Highest Level of Education', '1.3.1 Colleges Attended', '1.3.2 Partial College Credit', '1.4 Desired Degree Level', '1.5 Program of Interest', '1.5.1 Category of Interest', '1.6.1 Current Occupation', '1.6.2 Career Goal', '1.7 Desired Campus', '1.8 Date of birth', '2.1 Recommended Program', '2.2 Transfer Credits', '2.3 NAU + Transfer Credits', '2.4 Estimated Credits Remaining', '3.1 Start Date', '3.2 Program to Primary Plan', '3.3 Campus to Primary Plan', '3.4 Residency', '3.5 Link to Program in Catalog', '5.1 Student ID', '5.2 NAU User ID', '5.2.1 Primary Academic Program', '5.3 Credits Remaining Confirmed', '5.3.1 1221 Courses', '5.3.2 1224 Courses', '5.3.3 1227 Courses', '5.3.4 1228 Courses', '5.3.5 1231 Courses', '5.4 Advising Issue', '6.1 Area of Concern', '6.2 GPA', '6.3 Year in School', '6.4 Academic Standing', '6.5 Return Start Date', '7.1 Stop-out Reason', '9.2.1.1 Reason ReEngage Eligible', '9.4.1 Default', '9.8 New Lurkers', '9.9 Testing', 'Shared Credits', 'Shared Term']
#Making sure that none of the column names will break anything
for i in join_enrolled_v.columns:
    if i not in data_into_list:
        raise Exception(i + " is not an accepted custom field name")
#We want 0's not empty stuff on these
join_enrolled_v[["5.3.1 1221 Courses","5.3.2 1224 Courses","5.3.3 1227 Courses","5.3.4 1228 Courses","5.3.5 1231 Courses"]]=join_enrolled_v[["5.3.1 1221 Courses","5.3.2 1224 Courses","5.3.3 1227 Courses","5.3.4 1228 Courses","5.3.5 1231 Courses"]].fillna(0)
#Dropping the dupes so we only post once for each student
join_enrolled_v=join_enrolled_v.drop_duplicates(subset=["5.1 Student ID"])
#We only post what is in close, cuz otherwise it'll break
join_enrolled_v=join_enrolled_v[join_enrolled_v['id'].notna()]
#This is used for looping or smth
import_count = len(join_enrolled_v)
join_enrolled_v.to_csv(f'C:/Users/{username}/Downloads/jeV.csv',index=False,float_format='%.0f')
#join_enrolled_v
rem1224=remainX[remainX['5.3.2 1224 Courses']>0]
rem1227=remainX[remainX['5.3.3 1227 Courses']>0]
not1224=join_not_enrolled[join_not_enrolled['TERM']==1224]
not1227=join_not_enrolled[join_not_enrolled['TERM']==1227]
rem1224.to_csv('1224.csv',index=False)
rem1227.to_csv('1227.csv',index=False)
not1224.to_csv('not1224.csv',index=False)
not1227.to_csv('not1227.csv',index=False)
writing_to_workbook('wb.xlsx','wb.xlsx','1224.csv','Summer Enrolled','1227.csv','Fall Enrolled','not1224.csv','Summer Not Enrolled','not1227.csv','Fall Not Enrolled')
   


#Bulk assigning to what these should be based off of instructions, this shouln't ever have to be changed
#We only care about the ones that aren't in close, so we filter to custom.id is na since those are all the ones not in close
join_not_enrolled=join_not_enrolled[join_not_enrolled['custom.id'].isna()]
join_not_enrolled["Status"]="4.2 Application-ReAdmit-RegOfc"
join_not_enrolled["Full Name"]=join_not_enrolled["First_Name"]+" "+join_not_enrolled["Last_Name"]
join_not_enrolled["0.2 Assignee"]="James Allen"
join_not_enrolled["0.4.1 Lead Type"]="NAU Reg Office"
#Renaming to what it is in the original excel sheet since Vernon was insisten
join_not_enrolled=join_not_enrolled.rename(columns={"First_Name":"First Name","Last_Name":"Last Name","Primary_Acad_Prog":"Primary Academic Program"})
#Get rid of unneeeded columns
join_not_enrolled=join_not_enrolled[["EMPLID","Primary Academic Program","First Name","Last Name","Full Name","0.2 Assignee","0.4.1 Lead Type","Status","TERM"]]
#Get rid of dupes

mydate = datetime.now()
monthSTR=mydate.strftime("%B")
tday=datetime.today()
yday=tday-timedelta(days=1)
yday=datetime.strftime(yday,'%Y-%m-%d')
stuff=gsheet_to_df('Reengage Daily Performance Metrics 2022','list of stuff',client)
stuff=stuff[stuff['Date']==yday]
stuff=stuff.replace('4.*','4',regex=True)
stuff=stuff.replace('user_p8uqRNjqFh2i2rys71eEXQ6nomGcoexeXMfIyqW6FAL','Erik Van Contant')
stuff=stuff.replace('user_OZvsiWUHOkML9nxA02UOmo1MZ4pcPoLJfmtaSfImQAi','Vernon Eby')
stuff=stuff.replace('user_kxSWNpkEmPNixGaRDUDfWRdGjqxSeOA0Wt5IweBwDug','Vernon Rucker')
stuff=stuff.replace('user_HNVyAay5rLqRx88UGuAJOZI5fV6UM90DZpxnJdO0TFO','Crystal Merritte')
stuff=stuff.replace('user_hxH1Ombvp2IIjyalCPW92Zyd0qrYtQi4jBPPOHkbQv2','Ravenia Gant')
grouped_multiple = stuff.groupby(['GS', 'Status']).count()
grouped_multiple = grouped_multiple.reset_index()
ones=grouped_multiple[grouped_multiple['Status']=='1.0 ReEngage Qualified']
fours=grouped_multiple[grouped_multiple['Status']=='4']
current_day=gsheet_to_df('Reengage Daily Performance Metrics 2022','Current Day',client)
current_day=pd.merge(current_day,ones,left_on='User',right_on='GS',how='left')
current_day['Qualify Calls']=current_day['Date_y']
current_day=pd.merge(current_day,fours,left_on='User',right_on='GS',how='left')
current_day['Apps']=current_day['Date']
current_day=current_day.iloc[:, : 10]
current_day=current_day.rename(columns={"Date_x":"Date"})
try:
    month=gsheet_to_df('Reengage Daily Performance Metrics 2022',f'{monthSTR} 22',client)
except:
    print(traceback.format_exc())
    update=client.open("Reengage Daily Performance Metrics 2022")
    worksheet = update.add_worksheet(title=monthSTR+" 22", rows = 160, cols=12)
    month=gsheet_to_df('Reengage Daily Performance Metrics 2022',f'{monthSTR} 22',client,1)
current_day=current_day.fillna(0)
df_to_sheet('Reengage Daily Performance Metrics 2022','Current Day',current_day,client,1)
current_day=current_day.append(month)
current_day=current_day.fillna(0)
df_to_sheet('Reengage Daily Performance Metrics 2022',f'{monthSTR} 22',current_day,client,1)

   
#APIError: {'code': 400, 'message': 'Invalid requests[0].addSheet:
# A sheet with the name "August 22" already exists. Please enter
# another name.', 'status': 'INVALID_ARGUMENT'}

#Write the 3 remaining dataframes to csv's
remainX.to_csv(f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/Existing Students Course Count Updates/{today} Existing Course Count Upload.csv",index=False)
lost.to_csv(f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/Enrolled Not In Close/{today} Enrolled Not In Close.csv",index=False)
join_not_enrolled.to_csv(f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/New Students - Not Enrolled/{today} New Not Enrolled.csv",index=False)

#Enrolled, Not enrolled, Close, Existing Students, Active Students No Courses
#Write all these to the workbook at the location(v long)
writing_to_workbook(f"C:/Users/{username}/OneDrive/Desktop/{today} Enrollment Groups.xlsx",
                    f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/Enrollment Groups Workbooks - Complete/{today} Enrollment Groups.xlsx",enrolled,"Enrolled",not_enrolled,"Not Enrolled",close,"Close",
                    f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/Existing Students Course Count Updates/{today} Existing Course Count Upload.csv","Existing Students",
                    f"C:/Users/{username}/OneDrive/Desktop/Box Sync/Northern Arizona/Reporting/Uploads to close io/Enrollment Groups Workbook/New Students - Not Enrolled/{today} New Not Enrolled.csv","Active Students No Courses",
                    f'C:/Users/{username}/quarter.csv',"New Enrolled This Quarter",
                    f'C:/Users/{username}/quarter2.csv',"Existing Enrolled",
                    f'C:/Users/{username}/not.csv',"New Not Enrolled This Quarter",
                    f'C:/Users/{username}/not2.csv',"Existing Not Enrolled")
print("Workbook written")  
#TODOOO - CHange path name - /OneDrive/Desktop
#Webhook to send data to zapier to send the registration slack
data={'dropped':dfDropped2[['EMPLID','First_Name','Last_Name']].to_string(),'added':dfAdded2[['EMPLID','First_Name','Last_Name']].to_string(),'lost':lost.to_string(),'reg':("SU22 Registration including Online Ed = "+str(len1224)+", w/o= " + str(len(len1224O))+' '+"FA22 Registration including Online Ed = "+str(len1227)+", w/o= " + str(len(len1227O))+' '+"There are " + str(len(lost))+" rows in the Enrolled Not In Close workbook")}
r=requests.post('https://hooks.zapier.com/hooks/catch/2332311/bzkuxmo/',data=json.dumps(data),headers={'Content-Type':'application/json'})

#pool=db_c.database_connection()
#df.to_sql('users',con=pool,index=False)


#Clearing Automation 2
#We read the Important Info tab that we downloaded from the clearing checklist as a dataframe
info_df=pd.read_csv(f'C:/Users/{username}/Downloads/todo/Important Info {today}.csv')
#We group these cuz we just want counts for them, the number of classes for ex and the number of to do list items for todo
ex=enrolled_df.groupby(["EMPLID"]).count()
todo_df=df_merge.groupby(['EMPLID']).count()
#Setting the numbers
info_df=pd.merge(info_df,ex, left_on='ID', right_on='EMPLID', how='left')
info_df['Total Registered Course Count']=info_df['First_Name']
info_df=pd.merge(info_df,todo_df, left_on='ID', right_on='EMPLID', how='left')
info_df['Important To-Do List Items']=info_df['FIRST_NAME']
info_df=pd.merge(info_df,remainX, left_on='ID', right_on='5.1 Student ID', how='left')
info_df['Spring 2022 Courses']=info_df['5.3.1 1221 Courses']
info_df['Summer 2022 Courses']=info_df['5.3.2 1224 Courses']
info_df['Fall 2022 Courses']=info_df['5.3.3 1227 Courses']
info_df['Winter 2022 Courses']=info_df['5.3.4 1228 Courses']
info_df['Spring 2023 Courses']=info_df['5.3.5 1231 Courses']
#Theres gonna be a billion columns in this, but we know we only want like the first 25 or whatever so we get rid of everything else
info_df=info_df.iloc[:,:30]
info_df=info_df.drop(['status_label','Lead Type'],axis=1)
#We want to add lead type to this too, this is the easiest way to do so even though Im p sure we could grab it from somewhere else without having to do this
close_dfZ=close_df[['custom.5.1 Student ID','custom.0.4.1 Lead Type','custom.3.1 Start Date','status_label','custom.PS Clearing']]
info_df=pd.merge(info_df,close_dfZ,left_on='ID',right_on='custom.5.1 Student ID',how='left')
info_df['Target Start Date']=info_df['custom.3.1 Start Date']
info_df['Link To Checklist']=info_df['custom.PS Clearing']                    
info_df=info_df.drop(['custom.5.1 Student ID','custom.3.1 Start Date','custom.PS Clearing','First_Name','Last_Name'],axis=1)
info_df=info_df.rename(columns={'custom.0.4.1 Lead Type':'Lead Type'})
#These next three lines are just experimenting with dynamically changing fafsa based on the datasheet, but it loses a lot of info so idk if its useful
#faf=fafsa_df[fafsa_df['AID_YEAR']==2021]
#info_df2=pd.merge(info_df,faf,left_on='ID',right_on='PERS_EMPLID',how='left',indicator='FAFSA')
#info_df2.to_csv(f'C:/Users/GS/Downloads/todo/info2 {today}.csv')
Drop_Reason=info_df.pop("Drop Reason")
GS_Action=info_df.pop("GS Action")
Forecast=info_df.pop("Forecast")
info_df.insert(25, "Drop Reason", Drop_Reason )
info_df.insert(26, "GS Action", GS_Action )
info_df.insert(27, "Forecast", Forecast )
info_df.to_csv(f'C:/Users/{username}/Downloads/todo/info {today}.csv')
#Posting the thing back to ccr
df_to_sheet('Clearing Checklist Rest','Important Info',info_df,client,1)
closedfOMEGA=pd.merge(close_df,remainX,left_on='custom.5.1 Student ID',right_on='5.1 Student ID',how='left')
closedfOMEGA=closedfOMEGA[['status_label','display_name','custom.0.2 Assignee.name','custom.0.4.1 Lead Type','custom.3.1 Start Date','custom.3.2 Program to Primary Plan','custom.3.3 Campus to Primary Plan','custom.5.1 Student ID', '5.2.1 Primary Academic Program','5.3.1 1221 Courses','5.3.2 1224 Courses','5.3.3 1227 Courses','5.3.4 1228 Courses','5.3.5 1231 Courses']]
closedfOMEGA=closedfOMEGA.rename(columns={'display_name':'Name','custom.0.2 Assignee.name':'GS','custom.0.4.1 Lead Type':'Lead Type','custom.3.1 Start Date':'Start Date','custom.3.2 Program to Primary Plan':'Program','custom.3.3 Campus to Primary Plan':'Campus','custom.5.1 Student ID':'Student ID'})
close81=closedfOMEGA[closedfOMEGA['status_label']=='8.1 Graduation Pending']
close921=closedfOMEGA[closedfOMEGA['status_label']=='9.2.1 Ineligible ReEngage Program']
df_to_sheet('Clearing Checklist Rest','8.1',close81,client,1)
df_to_sheet('Clearing Checklist Rest','9.2.1',close921,client,1)


#Close auto upload stuff
reader = csv.DictReader(open(f'C:/Users/{username}/Downloads/jeV.csv'))
headers = reader.fieldnames
#We're really only expecting id, we dont want to mess with anything but the custom fields
expected_headers = ('company','url','status','contact','title','email','phone','mobile_phone','fax','address','address_1','address_2','city','state','zip','country','id')
# Remove trailing empty column headers
while not len(headers[-1].strip()):
    del headers[-1]
# Check for duplicated column names
if len(set(headers)) != len(headers):
    raise Exception('Cannot have duplicate column header names')
# Check for duplicates after normalization
def slugify(str, separator='_'):
    str = unidecode.unidecode(str).lower().strip()
    return re.sub(r'\W+', separator, str).strip(separator)
normalized_headers = [slugify(col) for col in headers]
if len(set(normalized_headers)) != len(normalized_headers):
    raise Exception('After column header names were normalized there were duplicate column header names')
# build a map of header names -> index in actual header row
header_indices = {col: i for (i, col) in enumerate(normalized_headers)}  # normalized columns as keys
header_indices.update({col: i for (i, col) in enumerate(headers)})  # add in original column names as keys
expected_headers = [col for col in normalized_headers if col in expected_headers]
custom_headers = list(set(normalized_headers) - set(expected_headers))  # non-recognized fields in slug-ed format
custom_headers = [headers[header_indices[normalized_col]]for normalized_col in custom_headers]
print("Recognized these column names:")
print(expected_headers)
if len(custom_headers):
    print("The following columns will be imported as custom fields:")
    print(custom_headers)
#Turning the rows into json stuff that we can put
#We could extremely easily change this to just read the data straight from the previous data frame
#It would be more efficient too, but we dont do that cuz the pros of being able to easily
#Examine and confirm the data outweighs the incredibly minor (sub .02 second) speedup
#and the formatting is nicer this way without having to finagle out decimals
#but still the code will be posted just in case
"""def lead_df(row):
    #print(row)
    leadID=row['id']
    lead = {'custom': {}}
    for field in custom_headers:
        if field in row:
            if not pd.isnull(row[field]):
                lead['custom'][field] = row[field]
    return lead,leadID
for i, row in df.iterrows():
    lead_df(row)
    lead,leadID = lead_df(row)
    if not lead:
        continue
    unique_leads[leadID] = lead"""
unique_leads = {}
def lead_from_row(row):
    row = {column_name: column_value.strip() for column_name, column_value in row.items()}  # strip unnecessary white spaces
    # check if the row isn't empty
    has_data = {column_name: column_value for column_name, column_value in row.items()
        if column_value
    }
    if not has_data:
        return None
    leadID=row['id']
    lead = {'custom': {}}
    for field in custom_headers:
        if field in row:
            if row[field]!="":
                lead['custom'][field] = row[field]
    return lead,leadID
for i, row in enumerate(reader):
    lead,leadID = lead_from_row(row)
    if not lead:
        continue
    unique_leads[leadID] = lead
#Giving a chance to review the data and make sure it isnt messed up
print(f'Found {len(unique_leads)} leads (grouped by leadID) from {import_count} contacts.')
print('Here is a sample lead (last row):')
print(json.dumps(unique_leads[leadID], indent=4))
"""This was just a pause to allow to review, unneeded aorn and removing it allows for completely total automation"""
tic = time.perf_counter()
#Uploading everything
for key,val in unique_leads.items():
    try:
        api.put(f'lead/{key}',val)
    except:
        print(f"ERROR, {key} failed to upload")
toc = time.perf_counter()
tocY= time.perf_counter()
#Timed
print(f"Uploaded the data in {toc - tic:0.3f} seconds")
print(f"Completed the automation in {tocY - ticY:0.3f} seconds")
print("All Done :)")




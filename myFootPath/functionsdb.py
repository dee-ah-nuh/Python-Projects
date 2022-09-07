from apiclient import discovery, errors
from httplib2 import Http
from oauth2client import file, tools
from oauth2client import  client as c
#from googleapiclient import errors
from datetime import datetime, timedelta
today=datetime.today().strftime('%Y-%m-%d')
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from googleapiclient.discovery import build
# from google_auth_oauthlib.flow import InstalledAppFlow #May be unneeded
# from google.auth.transport.requests import Request
from oauth2client.service_account import ServiceAccountCredentials
import re
import pandas as pd
#import numpy as np
import csv
import requests
import glob
import os.path
import time
import os
username=os.getlogin()
import gspread
import base64
#import email #may be unneeded
import webbrowser
#import csv
import json
#import closeio_api
import unidecode
from closeio_api import Client as CloseIO_API
#from requests.exceptions import ConnectionError
from email.mime.text import MIMEText
#import xlsxwriter
import traceback
import warnings as warning
import database_connector as db_c
#import Start_and_retain

config = json.load(open('config.json'))
ticY = time.perf_counter()

def get_emails(config['credentials_file_path4'],config['clientsecret_file_path4']):    
    SCOPES = ['https://mail.google.com/']    
    store = file.Storage(config['credentials_file_path4'])
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(clientsecret_file_path, SCOPES)
        credentials = tools.run_flow(flow, store)
    gmail = build('gmail', 'v1', credentials=credentials)
    result = gmail.users().messages().list(maxResults=25,userId='me').execute()
    messages = result.get('messages')  
 
    for msg in messages:
        txt = gmail.users().messages().get(userId='me', id=msg['id']).execute()      
        try:
            payload = txt['payload']
            headers = payload['headers']
            for d in headers:
                if d['name'] == 'Subject':
                   if d['value']=="Your ReEngage NAU export is ready":  
                        
                        body=base64.urlsafe_b64decode(txt.get("payload").get("body").get("data").encode("ASCII")).decode("utf-8")
                        listX=body.split()
                        for i in listX:
                            if i[0:4]=='http':
                                webbrowser.open(i)
                                return gmail
        except:
            print(traceback.format_exc())
    return gmail

#This function initializes the api services for google drive, sheets, and gspread
#It returns these as 3 credential's which will allow usage of these api's when passes
def get_api_services(credentials_file_path, clientsecret_file_path):
    # define credentials and client secret file paths
    #credentials_file_path = './credentials2/credentials.json'
    #clientsecret_file_path = './credentials2/client_secret.json'
   
    # define scope
    SCOPE = 'https://www.googleapis.com/auth/drive'

    # define store
    store = file.Storage(credentials_file_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = c.flow_from_clientsecrets(clientsecret_file_path, SCOPE)
        credentials = tools.run_flow(flow, store)

    # define API service
    http = credentials.authorize(Http())
    drive = discovery.build('drive', 'v3', http=http)
    sheets = discovery.build('sheets', 'v4', credentials=credentials)
    #Initialize gspread
    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
         "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    credentials = ServiceAccountCredentials.from_json_keyfile_name('./credentials3/client_secret.json', scope)
    client = gspread.authorize(credentials)
    #Return the allowances
    return drive, sheets, client
api = "api_26qZCo6enkWewoyUHQvbGd.4PZwmPyklAo0vfrk4yWGrd"
#This would allow us to get rid of the predownload from zapier and quickly and efficently query on close for only needed data
#But its still not faster than just predownloading and grabbing the data there lol so it remains dormant. AKa self building query driver
def grab(df):
    start=0
    end=200
    rows={}
    df=df.drop_duplicates(subset='EMPLID',keep='first')
    while start<len(df):
        #This grabs everything with a student id thats in the df youre giving to it
        #So it'd be a completely clean left join
        data={#we can only grab 200 things at a time with this call
        "_limit": 200,
        "query": {
            "queries": [
                {
                    "object_type": "lead",
                    "type": "object_type"
                },
                {
                    "queries": [{
                        "queries":[{
                                        "condition": {
                                            "mode": "full_words",
                                            "type": "term",
                                            "values": []
                                        },
                                        "field": {
                                            "custom_field_id": "lcf_8xBr0KeZQNBtZMOChusviGn94cJSoiLTaQVDHxNx7PC",
                                            "type": "custom_field"
                                        },
                                        "type": "field_condition"
                                    }],"type":"and"
                        }
                    ],
                    "type": "and"
                }
            ],
            "type": "and"
        },
        '_fields':{"lead":['id','custom']},
        "sort": []
        }
        #Here we grab in range of 200 len
        for key in df[start:end]['EMPLID']:
            data['query']['queries'][1]['queries'][0]['queries'][0]['condition']['values'].append(str(key))
        resp=api.post('data/search',data=data)
        for lead in resp['data']:
                #we only care about leadid and custom fields here
            rows[lead['id']]=lead['custom']
        #After we use the range we update it to be right
        start=end
        end+=200
    #we then transform the grabbed data into a dataframe with the stipulations we require
    df2=pd.DataFrame.from_dict(rows,orient='index')
    df2.reset_index(inplace=True)
    df2 = df2.rename(columns = {'index':'id'})
    print(df2.dtypes)
    return df2

#This writes to csv a excel file at a given sheet and then at the end saves the wb as endname
#It can take any number of csv,sheet arguments as long as you include both and theyre all being written to the same sheet
def writing_to_workbook(wbname,endname,*args):    
    #csv,sheet
    #We don't always know if the workbook we're tryna get exists or not, so we try to grab it and if that doesnt work we make it
    try:
        wb=load_workbook(wbname)
    except:
        wb=Workbook(wbname)
    #These are the lists of sheets and csv's after they get into dataframes
    #We make these as lists to make accessing easier
    listDF=[]
    listWS=[]
    listWSName=[]
    for i in range(0,len(args)):
        #The even numbers (aka the 1st arguments(0 focused is weird like that))
        #These are the csv's, so we turn them into dataframes and then append to the list of dataframes
        if i%2==0:
            listDF.append(pd.read_csv(args[i],header=None,index_col=None))
        #The odd numbers (aka the 2nd args)
        #These are the worksheets that we want to be writing to
        else:
            #Same thing as with the workbooks, we dont know if these exist already or not so we try grabbing it and if that doesnt work we make it
            try:
                listWS.append(wb[args[i]])
                listWSName.append(args[i])
            except:
                listWS.append(wb.create_sheet(args[i]))
                listWSName.append(args[i])
    #The briliance of the lists is that for any index the pairs are matched up
    #So for listDF[1], that will be the csv argmunet half of the csv,sheet pair, and so we can just use listWS[1] to get that too
    #We know these are the same length unless the user messed up the arguments, and if they did then it'll error which is good
    #Otherwise we just go through the full lengths
    for i in range(0,len(listWS)):
        #We're assigning the value of listWS[i] to a variable because we have to use append later and its the pandas append, not the list append
        #It matters and broke it when I tried to directly do listWS[i].append(r)
        x=listWS[i]
        print(x)
        #So here x is the tab in the excel sheet we want to append to
        #And r is the rth row in the dataframe listDF[i]
        try:
            for r in dataframe_to_rows(listDF[i],index=False,header=None):
                x.append(r)
        except:
            print(traceback.format_exc())
            with pd.ExcelWriter(wbname, mode = "a", engine = "openpyxl", if_sheet_exists = 'replace') as writer:
              listDF[i].to_excel(writer, sheet_name = listWSName[i], index = False)            
    wb.save(endname)


#This returns the most recently created file at a location that matches regex criteria
#Path is the location up to and including the file name, and is generally going to be where the regex is
#Extension is just the file type, nothing too crazy here
def recent_file(path,extension,location):
    #Assign the name
    files = glob.glob(path + extension)
    #Get the path for the location'th most recently created thing matching
    max_file=sorted(files,key=os.path.getctime)[-location]
    #Return that as a text string
    return max_file

#This writes a dataframe to a google sheet
#Wbname is the sheet name, sheetname is the tab name
#Df is the data frame, gsrpead is the gspread authorization (usually going to be client)
def df_to_sheet(wbname,sheetname,df,gspread,*arg):
    #Open shet
    df=df.fillna('')
    wb=gspread.open(wbname)
    #Open tab
    ws=wb.worksheet(sheetname)
    if arg:
        if arg[0]==1:
            ws.clear()
        if arg[0]==2:
            ws.append_rows(df.values.tolist())
            return
    #Write the values
    ws.update([df.columns.values.tolist()] + df.values.tolist())

#This is so I don't have a billion versions of the same exact process in the doing the dropped and added stuff
#Currently unused
def cleanMerge(merge1,merge2,*args):  
    df=pd.merge(merge1,merge2,on="EMPLID",how='left')
    df=df[df["First_Name_y"].isna()]
    df=df.drop_duplicates(subset=["EMPLID"],keep='first')
    if args:
        df=df.dropna(axis=1,how='all')
    return df

def create_message(sender, to, subject, message_text):
  """Create a message for an email.
  Args:
    sender: Email address of the sender.
    to: Email address of the receiver.
    subject: The subject of the email message.
    message_text: The text of the email message.
  Returns:
    An object containing a base64url encoded email object."""
  message = MIMEText(message_text)
  message['to'] = to
  message['from'] = sender
  message['subject'] = subject
  raw_message = base64.urlsafe_b64encode(message.as_string().encode("utf-8"))
  return {'raw': raw_message.decode("utf-8")}

def send_message(service, user_id, message):
  """Send an email message.
  Args:
    service: Authorized Gmail API service instance.
    user_id: User's email address. The special value "me" can be used to indicate the authenticated user.
    message: Message to be sent.
  Returns:
    Sent Message."""
  message = (service.users().messages().send(userId=user_id, body=message).execute())
  return message

###THis downloads a google sheet tab to csv given google api (here sheets), the sheet id, and the sheet name you want
def download_sheet_to_csv(sheets_instance, spreadsheet_id,sheet_name):
    #We grab that sheet
    result = sheets_instance.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=sheet_name).execute()
    today=datetime.today().strftime('%Y-%m-%d')
    #This sets the download loacation
    output_file = f'C:/Users/{username}/Downloads/todo/{sheet_name} {today}.csv'
    #We iterate across the result1 and write all its values into the csv
    with open(output_file, 'w',newline='') as f:
        writer = csv.writer(f)
        writer.writerows(result.get('values'))
    f.close()
    #We let us know that its done
    print(f'Successfully downloaded {sheet_name} {today}.csv')

#THis gets the id for any google sheet given a drive api and its name
def get_spreadsheet_id(api_service,spreadsheet_name):
    results = []
    page_token = None
    while True:
        try:
            param = {'q': 'mimeType="application/vnd.google-apps.spreadsheet"'}
            if page_token:
                param['pageToken'] = page_token
            files = api_service.files().list(**param).execute()
            results.extend(files.get('files'))
            page_token = files.get('nextPageToken')
            if not page_token:
                break
        except errors.HttpError as error:
            print(traceback.format_exc())
            print(f'An error has occurred: {error}')
            break
    spreadsheet_id = [result.get('id') for result in results if result.get('name') == spreadsheet_name][0]
    return spreadsheet_id

def gsheet_to_df(book,sheet,gspread,*arg):
    wb=gspread.open(book)
    ws=wb.worksheet(sheet)
    data=ws.get_all_values()
    if not arg:
        headers=data.pop(0)
        df=pd.DataFrame(data,columns=headers)
        return df
    else:
        df=pd.DataFrame(data)
        return df

